from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from .models import Institution, Course, TrainingCourses, TrainingLead
from django.contrib.auth.views import login_required
from openpyxl import Workbook
import openpyxl
from django.http import HttpResponse


# Create your views here.
def loginview(request):
    msg = ''
    #try:
    #    print(request.GET['next'])
    #except Exception as e:
    #    print(str(e))
    if request.method == 'POST':
        username = request.POST.get('username')
        print('aaaaaaaaaaaaa', username)
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user is None:
            msg = 'Invalid Credentials!!'
        else:
            login(request, user)
            return redirect('dashboard')
    return render(request, 'login.html', {'msg': msg})


def register_view(request):
    return render(request, 'register.html', {})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='/master/')
def dashboard_view(request):
    return render(request, 'dashboard.html', {})


@login_required(login_url='/master/')
def datatable_view(request):
    institution = Institution.objects.all()
    return render(request, 'data-table.html', {'institution': institution})


@login_required(login_url='/master/')
def data_delete_view(request,id):
    institute = Institution.objects.get(id=id)
    institute.delete()
    return redirect('datatable')


@login_required(login_url='/master/')
def institution_add_view(request):
    if request.method == 'POST':
        InstitutionName = request.POST.get('InstitutionName')
        print(InstitutionName)
        InstitutionUniversity = request.POST.get('InstitutionUniversity')
        InstitutionContact = request.POST.get('InstitutionContact')
        InstitutionEmail = request.POST.get('InstitutionEmail')
        InstitutionWebsite = request.POST.get('InstitutionWebsite')
        InstitutionAddress = request.POST.get('InstitutionAddress')

        institution = Institution()
        institution.name = InstitutionName
        institution.university = InstitutionUniversity
        institution.contact_number = InstitutionContact
        institution.email = InstitutionEmail
        institution.website = InstitutionWebsite
        institution.address = InstitutionAddress
        institution.save()
        return redirect('datatable')
    return render(request, 'institution_add.html', {})


@login_required(login_url='/master/')
def institution_edit_view(request, id):
    institute = Institution.objects.get(id=id)
    if request.method == 'POST':
        InstitutionName = request.POST.get('InstitutionName')
        print(InstitutionName)
        InstitutionUniversity = request.POST.get('InstitutionUniversity')
        InstitutionContact = request.POST.get('InstitutionContact')
        InstitutionEmail = request.POST.get('InstitutionEmail')
        InstitutionWebsite = request.POST.get('InstitutionWebsite')
        InstitutionAddress = request.POST.get('InstitutionAddress')

        institute.name = InstitutionName
        institute.university = InstitutionUniversity
        institute.contact_number = InstitutionContact
        institute.email = InstitutionEmail
        institute.website = InstitutionWebsite
        institute.address = InstitutionAddress
        institute.save()
        return redirect('datatable')
    return render(request, 'institution_edit.html', {'institute': institute})


@login_required(login_url='/master/')
def cource_view(request):
    cource = Course.objects.all()
    return render(request, 'cource.html', {'cource': cource})


@login_required(login_url='/master/')
def cource_add_view(request):
    if request.method == "POST":
        courcename = request.POST.get('CourceName')
        courcetype = request.POST.get('CourceType')
        courcedescription = request.POST.get('CourceDescription')
        print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa", courcedescription)

        cource = Course()
        cource.name = courcename
        cource.type = courcetype
        cource.description = courcedescription
        cource.save()
        return redirect('cource_view')
    return render(request, 'cource_add.html', {})


@login_required(login_url='/master/')
def cource_edit_view(request,id):
    cource = Course.objects.get(id=id)
    if request.method == 'POST':
        courcename = request.POST.get('CourceName')
        courcetype = request.POST.get('CourceType')
        courcedescription = request.POST.get('CourceDescription')

        cource.name = courcename
        cource.type = courcetype
        cource.description = courcedescription
        cource.save()
        return redirect('cource_view')
    return render(request, 'cource_edit.html', {'cource': cource})


@login_required(login_url='/master/')
def cource_delete_view(request,id):
    cource = Course.objects.get(id=id)
    cource.delete()
    return redirect('cource_view')


@login_required(login_url='/master/')
def training_view(request):
    training = TrainingCourses.objects.all()
    return render(request, 'training_view.html', {'training': training})


@login_required(login_url='/master/')
def training_add_view(request):
    if request.method == 'POST':
        TrainingName = request.POST.get('TrainingName')
        TrainingFee = request.POST.get('TrainingFee')
        TrainingType = request.POST.get('TrainingType')
        TrainingDescription = request.POST.get('TrainingDescription')

        training = TrainingCourses()
        training.name=TrainingName
        training.type = TrainingType
        training.fee = TrainingFee
        training.description = TrainingDescription
        training.save()
        return redirect('training_view')
    return render(request, 'training_add.html', {})


@login_required(login_url='/master/')
def training_edit_view(request,id):
    training = TrainingCourses.objects.get(id=id)
    if request.method == 'POST':
        TrainingName = request.POST.get('TrainingName')
        TrainingFee = request.POST.get('TrainingFee')
        TrainingType = request.POST.get('TrainingType')
        TrainingDescription = request.POST.get('TrainingDescription')

        training.name = TrainingName
        training.type = TrainingType
        training.fee = TrainingFee
        training.description = TrainingDescription
        training.save()
        return redirect('training_view')
    return render(request, 'training_edit.html', {'training': training})


@login_required(login_url='/master/')
def training_delete_view(request,id):
    training = TrainingCourses.objects.get(id=id)
    training.delete()
    return redirect('training_view')


@login_required(login_url='/master/')
def traininglead_view(request):
    temp = request.POST.get('follow')
    print("aaaaaaaaaaaaa",temp)
    # f= datetime.date(2019,8,19)
    # t = datetime.date(2019, 8, 21)
    f = request.POST.get('from')
    t = request.POST.get('to')
    # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaa",data)
    if temp == "on":
        if request.user.is_superuser:
            if request.method=='POST':
                traininglead = TrainingLead.objects.filter(followup__range=(f, t))
            else:
                traininglead = TrainingLead.objects.all()
        else:
            if request.method == 'POST':
                traininglead = TrainingLead.objects.filter(followup__range=(f, t), created_by=request.user)
            else:
                traininglead = TrainingLead.objects.all(created_by=request.user)
    else:
        if request.user.is_superuser:
            if request.method=='POST':
                traininglead = TrainingLead.objects.filter(enquery_date__range=(f, t))
            else:
                traininglead = TrainingLead.objects.all()
        else:
            if request.method == 'POST':
                traininglead = TrainingLead.objects.filter(enquery_date__range=(f, t), created_by=request.user)
            else:
                traininglead = TrainingLead.objects.all(created_by=request.user)

    if request.method == 'POST' and "export" in request.POST:
        traininglead = request.POST.getlist('export')
        print(traininglead)
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Name"
        ws.cell(1, 1).font = ws.cell(1, 1).font.copy(bold=True)
        ws.cell(1, 2).value = "College"
        ws.cell(1, 2).font = ws.cell(1, 2).font.copy(bold=True)
        ws.cell(1, 3).value = "Course"
        ws.cell(1, 3).font = ws.cell(1, 3).font.copy(bold=True)
        ws.cell(1, 4).value = "Year of Passwout"
        ws.cell(1, 4).font = ws.cell(1, 4).font.copy(bold=True)
        ws.cell(1, 5).value = "Email"
        ws.cell(1, 5).font = ws.cell(1, 5).font.copy(bold=True)
        ws.cell(1, 6).value = "Mobile"
        ws.cell(1, 6).font = ws.cell(1, 6).font.copy(bold=True)
        ws.cell(1, 7).value = "Enquiry Date"
        ws.cell(1, 7).font = ws.cell(1, 7).font.copy(bold=True)
        ws.cell(1, 8).value = "Enquiry For"
        ws.cell(1, 8).font = ws.cell(1, 8).font.copy(bold=True)
        ws.cell(1, 9).value = "Remarks"
        ws.cell(1, 9).font = ws.cell(1, 9).font.copy(bold=True)
        ws.cell(1, 10).value = "Follow Up Date"
        ws.cell(1, 10).font = ws.cell(1, 10).font.copy(bold=True)
        ws.cell(1, 11).value = "Last Follow up Date"
        ws.cell(1, 11).font = ws.cell(1, 11).font.copy(bold=True)
        ws.cell(1, 12).value = "Status"
        ws.cell(1, 12).font = ws.cell(1, 12).font.copy(bold=True)
        k = 2
        for i in traininglead:
            j = 1
            lead = TrainingLead.objects.get(id=i)
            ws.cell(k, j).value = lead.name
            j += 1
            ws.cell(k, j).value = str(lead.institution)
            print(lead.institution)
            j += 1
            ws.cell(k, j).value = str(lead.course)
            j += 1
            ws.cell(k, j).value = lead.year_of_passout
            j += 1
            ws.cell(k, j).value = lead.email
            j += 1
            ws.cell(k, j).value = lead.mobile
            j += 1
            ws.cell(k, j).value = lead.enquery_date
            j += 1
            ws.cell(k, j).value = str(lead.enquery_for)
            j += 1
            ws.cell(k, j).value = lead.remarks
            j += 1
            ws.cell(k, j).value = lead.followup
            j += 1
            ws.cell(k, j).value = lead.lastfollowup
            j += 1
            ws.cell(k, j).value = lead.status
            k += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        #
        wb.save(response)
        return response

    return render(request, 'traininglead_view.html', {'traininglead': traininglead, 'temp': temp, 'f': f, 't': t})


@login_required(login_url='/master/')
def traininglead_add_view(request):
    institute = Institution.objects.all()
    cource = Course.objects.all()
    training = TrainingCourses.objects.all()

    if request.method == 'POST':
        TrainingleadName = request.POST.get('TrainingleadName')
        TrainingleadInstitute = request.POST.get('TrainingleadInstitute')
        TrainingleadCource = request.POST.get('TrainingleadCource')
        TrainingleadSemester = request.POST.get('TrainingleadSemester')
        TrainingleadYearofpass = request.POST.get('TrainingleadYearofpass')
        TrainingleadEmail = request.POST.get('TrainingleadEmail')
        TrainingleadMobile = request.POST.get('TrainingleadMobile')
        TrainingleadWhatsapp = request.POST.get('TrainingleadWhatsapp')
        TrainingleadEnqueryDate = request.POST.get('TrainingleadEnqueryDate')
        TrainingleadEnqueryFor = request.POST.get('TrainingleadEnqueryFor')
        TrainingleadCourceUpdate = request.POST.get('TrainingleadCourceUpdate')
        TrainingleadCourceFee = request.POST.get('TrainingleadCourceFee')
        if request.POST.get('TrainingleadIsJoined') is 'on':
            TrainingleadIsJoined = True
        else:
            TrainingleadIsJoined = False

        TrainingleadRemark = request.POST.get('TrainingleadRemark')


        traininglead = TrainingLead()

        traininglead.name = TrainingleadName
        traininglead.institution = Institution.objects.get(name=TrainingleadInstitute)
        traininglead.course = Course.objects.get(name=TrainingleadCource)
        traininglead.semester = TrainingleadSemester
        traininglead.year_of_passout = TrainingleadYearofpass
        traininglead.email = TrainingleadEmail
        traininglead.mobile = TrainingleadMobile
        traininglead.whatsapp_no = TrainingleadWhatsapp
        traininglead.enquery_date = TrainingleadEnqueryDate
        traininglead.enquery_for = TrainingCourses.objects.get(name=TrainingleadEnqueryFor)
        if TrainingleadCourceUpdate:
            traininglead.course_updated = TrainingCourses.objects.get(name=TrainingleadCourceUpdate)
        traininglead.course_fee = TrainingleadCourceFee
        traininglead.is_joined = TrainingleadIsJoined
        traininglead.remarks = TrainingleadRemark
        traininglead.created_by = request.user
        traininglead.save()

        return redirect('traininglead_view')

    return render(request, 'traininglead_add.html', {'institute': institute, 'cource': cource, 'training': training})

@login_required(login_url='/master/')
def traininglead_edit_view(request,id):
    traininglead = TrainingLead.objects.get(id=id)

    cource = Course.objects.all()
    training = TrainingCourses.objects.all()

    if request.method == 'POST':
        TrainingleadName = request.POST.get('TrainingleadName')
        TrainingleadInstitute = request.POST.get('TrainingleadInstitute')
        TrainingleadCource = request.POST.get('TrainingleadCource')
        TrainingleadYearofpass = request.POST.get('TrainingleadYearofpass')
        TrainingleadEmail = request.POST.get('TrainingleadEmail')
        TrainingleadMobile = request.POST.get('TrainingleadMobile')
        TrainingleadEnqueryDate = request.POST.get('TrainingleadEnqueryDate')
        TrainingleadEnqueryFor = request.POST.get('TrainingleadEnqueryFor')
        TrainingleadRemark = request.POST.get('TrainingleadRemark')

        print(request.POST)

        traininglead.name = TrainingleadName
        traininglead.institution = Institution.objects.get(name=TrainingleadInstitute)
        traininglead.course = Course.objects.get(name=TrainingleadCource)
        traininglead.year_of_passout = TrainingleadYearofpass
        traininglead.email = TrainingleadEmail
        traininglead.mobile = TrainingleadMobile
        traininglead.enquery_date = TrainingleadEnqueryDate
        traininglead.enquery_for = TrainingCourses.objects.get(name=TrainingleadEnqueryFor)
        traininglead.remarks = TrainingleadRemark
        traininglead.save()

        return redirect('traininglead_view')
    return render(request, 'traininglead_edit.html', {'traininglead': traininglead, 'cource': cource, 'training': training})

@login_required(login_url='/master/')
def traininglead_delete_view(request,id):
    traininglead = TrainingLead.objects.get(id=id)
    traininglead.delete()
    return redirect('traininglead_view')


@login_required(login_url='/master/')
def traininglead_followup_view(request, id):
    traininglead = TrainingLead.objects.get(id=id)
    if request.method == 'POST':

        followup = request.POST.get('TrainingleadFollowupDate')
        status = request.POST.get('TrainingleadStatus')
        remark = request.POST.get('TrainingleadRemark')
        lastfollow = request.POST.get('TrainingleadLastFollowupDate')

        traininglead.followup = followup
        traininglead.status = status
        traininglead.remark = remark
        traininglead.lastfollowup = lastfollow
        traininglead.save()
        return redirect('traininglead_view')

    return render(request, 'traininglead_followup.html', {'traininglead': traininglead})
